import openpyxl
import subprocess

from constants import names

file_path = "Dekabr 2024.xlsx"
#subprocess.Popen(file_path, shell=True)
sheet_name = "2024 Dekabr (30)"
workbook = openpyxl.load_workbook(file_path, data_only=True, keep_vba=False, keep_links=False)
#sheet = workbook.active
sheet = workbook[sheet_name]

def read_excel_data():
    
    data = {}

    for col in range(1, 17):
        for row in range(9, 24):
            actual_row = row - 8
            cell = sheet.cell(row=row, column=col*2)
            if cell.value:
                if names[col-1] not in data:
                    data[names[col-1]] = {}

                data[names[col-1]][actual_row] = cell.value
    
    return data

# def save_new_record(doctor, amount):
#     if doctor in data:
#         new_row = len(data[doctor]) + 1
#     else:
#         data[doctor] = {}
#         new_row = 1
    
#     data[doctor][new_row] = amount

#     col = (names.index(doctor) + 1)*2
    
#     sheet.cell(row=9 + new_row - 1, column=col).value = amount
#     workbook.save(file_path)